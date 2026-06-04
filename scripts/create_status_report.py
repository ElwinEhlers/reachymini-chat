"""Generate an Excel status report for the Reachy Chat project."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

wb = Workbook()

# ── Sheet 1: Project Overview ──
ws1 = wb.active
ws1.title = "Projektübersicht"

header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2E3A59", end_color="2E3A59", fill_type="solid")
section_font = Font(bold=True, size=11, color="1A1A2E")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_cell(ws, row, col, border=True):
    cell = ws.cell(row=row, column=col)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if border:
        cell.border = thin_border
    return cell

# Title
ws1.merge_cells("A1:D1")
title_cell = ws1["A1"]
title_cell.value = "Reachy Mini Chat App — Projektstatus"
title_cell.font = Font(bold=True, size=16, color="1A1A2E")
title_cell.alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:D2")
ws1["A2"].value = f"Stand: {date.today().strftime('%d.%m.%Y')}"
ws1["A2"].font = Font(size=11, color="666666")
ws1["A2"].alignment = Alignment(horizontal="center")

# Column widths
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 15
ws1.column_dimensions["D"].width = 30

# Project info section
row = 4
ws1.merge_cells(f"A{row}:D{row}")
ws1.cell(row=row, column=1, value="Projektinformationen").font = section_font
ws1.cell(row=row, column=1).fill = section_fill

info = [
    ("Projektname", "Reachy Mini Chat App (reachy_chat)"),
    ("Beschreibung", "Browser-basierte Konversations-App für Reachy Mini mit lokalem LLM"),
    ("Arbeitsverzeichnis", "/home/sbin/reachy"),
    ("App-Verzeichnis", "/home/sbin/reachy/reachy_chat"),
    ("OS", "Ubuntu Desktop (Linux 6.17)"),
    ("Python", "3.12.3 (.venv via uv)"),
    ("SDK", "reachy-mini v1.5.1"),
    ("LLM", "Ollama qwen3:8b-q4km (lokal, localhost:11434)"),
    ("Web-UI", "http://localhost:8042"),
    ("Daemon", "uv run reachy-mini-daemon --sim (localhost:8000)"),
    ("Modus", "Simulation (MuJoCo) — Hardware noch ausstehend"),
]

for key, val in info:
    row += 1
    style_cell(ws1, row, 1).value = key
    ws1.cell(row=row, column=1).font = Font(bold=True)
    ws1.merge_cells(f"B{row}:D{row}")
    style_cell(ws1, row, 2).value = val

# ── Sheet 2: Completed Tasks ──
ws2 = wb.create_sheet("Erledigte Aufgaben")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 55
ws2.column_dimensions["D"].width = 12

row = 1
headers = ["#", "Aufgabe", "Details", "Status"]
for col, h in enumerate(headers, 1):
    ws2.cell(row=row, column=col, value=h)
style_header_row(ws2, 1, 4)

tasks = [
    (1, "Ollama installieren & Modell laden",
     "Ollama installiert, llama3.1:8b Modell heruntergeladen. "
     "Erreichbar unter localhost:11434. Unterstützt OpenAI-kompatibles Tool-Calling.",
     "Erledigt"),
    (2, "App mit Conversation-Template erstellen",
     "reachy-mini-app-assistant create --template conversation reachy_chat. "
     "Erstellt vollständiges Projekt mit Tool-System (dance, emotions, sweep_look etc.).",
     "Erledigt"),
    (3, "Code aufräumen & anpassen",
     "Entfernt: openai_realtime.py, console.py, gradio_personality.py, head_wobbler.py, "
     "headless_personality*.py. pyproject.toml: Gradio/fastrtc/aiortc entfernt. "
     "config.py: OpenAI-Referenzen durch OLLAMA_BASE_URL/OLLAMA_MODEL ersetzt.",
     "Erledigt"),
    (4, "ollama_chat.py implementieren",
     "Neues LLM-Backend: AsyncOpenAI-Client gegen Ollama, Chat-Completion mit "
     "Tool-Calling, SSE-Streaming, Message-History, System-Prompt aus Profil.",
     "Erledigt"),
    (5, "main.py umschreiben",
     "Vereinfachter Einstiegspunkt: MovementManager + OllamaChatHandler + "
     "FastAPI-Routen (POST /chat, GET /status, GET /history, POST /clear). "
     "ReachyChat-Klasse als ReachyMiniApp Entry-Point.",
     "Erledigt"),
    (6, "Browser Chat-UI erstellen",
     "Dark Theme Chat-Interface: static/index.html, main.js, style.css. "
     "Features: SSE-Streaming, Typing-Indikator, Tool-Call-Anzeige, "
     "Status-Badge, Auto-Resize Textarea, Enter zum Senden.",
     "Erledigt"),
    (7, "Profil konfigurieren & End-to-End-Test",
     "instructions.txt und tools.txt angepasst. Tools aktiviert: dance, stop_dance, "
     "play_emotion, stop_emotion, move_head, sweep_look. "
     "Test erfolgreich: Chat-Anfrage → Emotion ausgelöst → Text gestreamt.",
     "Erledigt"),
    (8, "Modell gewechselt: qwen3:8b-q4km",
     "llama3.1:8b durch qwen3:8b-q4km ersetzt. Modell läuft stabil mit Tool-Calling, "
     "gute Mehrsprachigkeit (DE/EN). think-Modus entfernt (verursachte leere Antworten). "
     "Modell via 'ollama create' aus Blob erstellt. Konfiguration in .env.",
     "Erledigt"),
    (9, "Memory-System implementiert",
     "SQLite-Datenbank (WAL) unter memory/reachy.db. "
     "Module: database.py, persons.py, conversations.py, facts.py, context.py. "
     "GET /memory API-Endpunkt. Wird optional beim Start geladen. "
     "memory_status.html für Live-Ansicht der DB.",
     "Erledigt"),
    (10, "Sprach-System implementiert",
     "WebSocket-Endpoint /ws/voice. Push-to-Talk im Browser. "
     "STT: faster-whisper (small, Deutsch). TTS: Piper (de_DE-thorsten-high). "
     "VAD: RMS-basierte Stille-Erkennung. Pipeline: voice/pipeline.py. "
     "Bekanntes Problem: sd.InputStream bei 16kHz liefert Stille via PipeWire → "
     "Fix: Aufnahme mit nativer Rate (44kHz), Resampling auf 16kHz.",
     "Erledigt"),
    (11, "GitHub-Dokumentation erstellt",
     "README.md komplett neu geschrieben (Deutsch + englische Kurzfassung). "
     "Abschnitte: Voraussetzungen, Installation, Konfiguration, Starten, "
     "Verzeichnisstruktur, Architektur, Eigene Tools & Profile. "
     ".env.example aktualisiert (OpenAI entfernt, Ollama ergänzt). "
     "Gepusht auf https://github.com/ElwinEhlers/reachymini-chat.",
     "Erledigt"),
    (12, "Background-Tool-System & Task-Management",
     "BackgroundToolManager für async Task-Ausführung. "
     "ToolState-Enum: RUNNING, COMPLETED, FAILED, CANCELLED. "
     "System-Tools: task_status, task_cancel (automatisch geladen).",
     "Erledigt"),
    (13, "custom_tool.py Template im Profil",
     "Fertiges Template für eigene Tools im _reachy_chat_locked_profile. "
     "Direkt editierbar ohne neue Dateien anlegen zu müssen.",
     "Erledigt"),
]

for t in tasks:
    row += 1
    for col, val in enumerate(t, 1):
        cell = style_cell(ws2, row, col)
        cell.value = val
    ws2.cell(row=row, column=4).fill = done_fill

# ── Sheet 3: File Changes ──
ws3 = wb.create_sheet("Dateiänderungen")

ws3.column_dimensions["A"].width = 55
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 45

headers = ["Datei", "Aktion", "Beschreibung"]
for col, h in enumerate(headers, 1):
    ws3.cell(row=1, column=col, value=h)
style_header_row(ws3, 1, 3)

files = [
    ("src/reachy_chat/ollama_chat.py", "NEU", "LLM-Backend mit Ollama + Tool-Calling"),
    ("src/reachy_chat/main.py", "ERSETZT", "Vereinfachter Einstiegspunkt mit FastAPI-Routen"),
    ("src/reachy_chat/config.py", "GEÄNDERT", "OpenAI → Ollama Konfiguration"),
    ("src/reachy_chat/static/index.html", "ERSETZT", "Chat-Interface statt API-Key-Seite"),
    ("src/reachy_chat/static/main.js", "ERSETZT", "Chat-Logik mit SSE-Streaming"),
    ("src/reachy_chat/static/style.css", "ERSETZT", "Dark Theme Chat-UI"),
    ("profiles/.../instructions.txt", "GEÄNDERT", "System-Prompt für Reachy"),
    ("profiles/.../tools.txt", "GEÄNDERT", "6 Tools aktiviert"),
    ("pyproject.toml", "GEÄNDERT", "Gradio/fastrtc entfernt, Dependencies bereinigt"),
    ("src/reachy_chat/openai_realtime.py", "GELÖSCHT", "OpenAI Realtime WebSocket API"),
    ("src/reachy_chat/console.py", "GELÖSCHT", "Headless Audio-Stream"),
    ("src/reachy_chat/gradio_personality.py", "GELÖSCHT", "Gradio UI"),
    ("src/reachy_chat/audio/head_wobbler.py", "GELÖSCHT", "Audio-reaktive Kopfbewegung"),
    ("src/reachy_chat/headless_personality.py", "GELÖSCHT", "Headless Personality"),
    ("src/reachy_chat/headless_personality_ui.py", "GELÖSCHT", "Headless Personality UI"),
]

new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
changed_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
replaced_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

action_fills = {
    "NEU": new_fill,
    "GEÄNDERT": changed_fill,
    "GELÖSCHT": deleted_fill,
    "ERSETZT": replaced_fill,
}

for i, (path, action, desc) in enumerate(files, 2):
    style_cell(ws3, i, 1).value = path
    ws3.cell(row=i, column=1).font = Font(name="Consolas", size=10)
    cell = style_cell(ws3, i, 2)
    cell.value = action
    cell.fill = action_fills.get(action, PatternFill())
    cell.alignment = Alignment(horizontal="center", vertical="top")
    style_cell(ws3, i, 3).value = desc

# ── Sheet 4: Architecture ──
ws4 = wb.create_sheet("Architektur")

ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 55

headers = ["Komponente", "Details"]
for col, h in enumerate(headers, 1):
    ws4.cell(row=1, column=col, value=h)
style_header_row(ws4, 1, 2)

arch = [
    ("LLM", "Ollama (qwen3:8b-q4km) via OpenAI-kompatible API auf localhost:11434"),
    ("Chat-Backend", "ollama_chat.py: AsyncOpenAI Client, Streaming, Tool-Calling Loop (max 5 Runden)"),
    ("Web-Server", "FastAPI (via ReachyMiniApp.settings_app) auf Port 8042"),
    ("Chat-UI", "Vanilla HTML/JS/CSS, Dark Theme, SSE-Streaming"),
    ("Movement", "MovementManager: 100Hz Control-Loop, Primär-/Sekundärbewegungen"),
    ("Tools", "dance, stop_dance, play_emotion, stop_emotion, move_head, sweep_look, task_status, task_cancel"),
    ("Tool-System", "core_tools.py: Tool-Registry, Dispatch, Profil-basierte Konfiguration"),
    ("Profil", "_reachy_chat_locked_profile (instructions.txt + tools.txt + custom_tool.py)"),
    ("Memory", "SQLite WAL (memory/reachy.db): Personen, Fakten, Gesprächsverläufe — optional"),
    ("Voice", "WebSocket /ws/voice: faster-whisper STT + Piper TTS + VAD — Deutsch"),
    ("Robot SDK", "reachy-mini v1.5.1 (set_target, goto_target, Emotions-Library)"),
    ("Daemon", "reachy-mini-daemon --sim auf localhost:8000"),
]

for i, (comp, detail) in enumerate(arch, 2):
    style_cell(ws4, i, 1).value = comp
    ws4.cell(row=i, column=1).font = Font(bold=True)
    style_cell(ws4, i, 2).value = detail

# ── Sheet 5: Next Steps ──
ws5 = wb.create_sheet("Nächste Schritte")

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 12
ws5.column_dimensions["C"].width = 35
ws5.column_dimensions["D"].width = 50

headers = ["#", "Phase", "Aufgabe", "Details"]
for col, h in enumerate(headers, 1):
    ws5.cell(row=1, column=col, value=h)
style_header_row(ws5, 1, 4)

next_steps = [
    (1, "Phase 1", "Voice-Pipeline verifizieren",
     "PipeWire/sd.InputStream-Fix in vad.py nach App-Neustart bestätigen (Aufnahme bei 44kHz, Resample auf 16kHz)"),
    (2, "Phase 1", "Modell-Auswahl in UI",
     "Dropdown zum Wechseln des Ollama-Modells im Browser"),
    (3, "Phase 1", "Memory-Integration in Chat-Kontext",
     "Bekannte Personen und Fakten automatisch in System-Prompt einbetten"),
    (4, "Phase 2", "Hardware-Inbetriebnahme",
     "USB-Verbindung, udev-Regeln bereits konfiguriert, Daemon ohne --sim starten"),
    (5, "Phase 2", "Kamera-Integration",
     "USB-Kamera aktivieren, CameraWorker einbinden"),
    (6, "Phase 2", "Person-Detection Tool",
     "MediaPipe/YOLO für Personenerkennung im Raum"),
    (7, "Phase 2", "Gesichtserkennung",
     "Bekannte Personen (aus Memory) erkennen und namentlich ansprechen"),
    (8, "Phase 2", "Head-Tracking",
     "Roboter folgt Personen mit dem Kopf automatisch (head_tracking Tool)"),
]

for t in next_steps:
    row = t[0] + 1
    for col, val in enumerate(t, 1):
        cell = style_cell(ws5, row, col)
        cell.value = val
    phase_cell = ws5.cell(row=row, column=2)
    if "Phase 1" in str(t[1]):
        phase_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    else:
        phase_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# ── Sheet 6: App-Dokumentation ──
ws6 = wb.create_sheet("App-Dokumentation")

ws6.column_dimensions["A"].width = 25
ws6.column_dimensions["B"].width = 75

doc_section_font = Font(bold=True, size=12, color="FFFFFF")
doc_section_fill = PatternFill(start_color="2E3A59", end_color="2E3A59", fill_type="solid")
doc_subsection_font = Font(bold=True, size=11, color="1A1A2E")
doc_subsection_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
code_font = Font(name="Consolas", size=10)

def doc_section(ws, row, title):
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = doc_section_font
    cell.fill = doc_section_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    return row + 1

def doc_subsection(ws, row, title):
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = doc_subsection_font
    cell.fill = doc_subsection_fill
    cell.border = thin_border
    ws.cell(row=row, column=2).border = thin_border
    return row + 1

def doc_row(ws, row, label, value, is_code=False):
    c1 = style_cell(ws, row, 1)
    c1.value = label
    c1.font = Font(bold=True)
    c2 = style_cell(ws, row, 2)
    c2.value = value
    if is_code:
        c2.font = code_font
    return row + 1

def doc_text(ws, row, text):
    ws.merge_cells(f"A{row}:B{row}")
    cell = style_cell(ws, row, 1)
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row, column=2).border = thin_border
    return row + 1

r = 1

# ── Title ──
ws6.merge_cells("A1:B1")
ws6["A1"].value = "Reachy Chat App — Technische Dokumentation"
ws6["A1"].font = Font(bold=True, size=14, color="1A1A2E")
ws6["A1"].alignment = Alignment(horizontal="center")
r = 3

# ── 1. Quickstart ──
r = doc_section(ws6, r, "1. Quickstart — App starten")
r = doc_row(ws6, r, "Voraussetzungen", "Daemon muss laufen, Ollama muss laufen mit Modell")
r = doc_row(ws6, r, "Schritt 1: Daemon", "uv run reachy-mini-daemon --sim", True)
r = doc_row(ws6, r, "Schritt 2: App starten", "uv run reachy-chat", True)
r = doc_row(ws6, r, "Schritt 3: Browser", "http://localhost:8042")
r = doc_row(ws6, r, "Daemon-Dashboard", "http://localhost:8000")
r = doc_row(ws6, r, "Ollama prüfen", "curl http://localhost:11434/v1/models", True)
r += 1

# ── 2. Projektstruktur ──
r = doc_section(ws6, r, "2. Projektstruktur")
structure = [
    ("reachy_chat/", "Projekt-Root"),
    ("  src/reachy_chat/", "Hauptpaket"),
    ("    main.py", "Einstiegspunkt: ReachyChat-Klasse, FastAPI-Routen, run()-Funktion"),
    ("    ollama_chat.py", "LLM-Backend: OllamaChatHandler mit Tool-Calling und SSE-Streaming"),
    ("    config.py", "Konfiguration: OLLAMA_BASE_URL, OLLAMA_MODEL, Profil-Verwaltung"),
    ("    moves.py", "MovementManager: 100Hz Control-Loop, Breathing, Move-Queue"),
    ("    prompts.py", "System-Prompt laden aus Profil, [includes] expandieren"),
    ("    utils.py", "CLI-Args, Logger-Setup, Vision-Handling"),
    ("    camera_worker.py", "Kamera-Thread (Phase 2)"),
    ("    dance_emotion_moves.py", "Tanz- und Emotions-Bewegungen (GotoQueueMove etc.)"),
    ("    static/", "Browser-UI Dateien"),
    ("      index.html", "Chat-Interface HTML"),
    ("      main.js", "Chat-Logik: SSE-Streaming, Nachrichten-Rendering"),
    ("      style.css", "Dark Theme Styling"),
    ("    tools/", "Tool-System"),
    ("      core_tools.py", "Tool-Basisklasse, Registry, Dispatch, Profil-Loader"),
    ("      dance.py", "Tanz-Tool: Tanz starten (zufällig oder benannt)"),
    ("      stop_dance.py", "Tanz stoppen"),
    ("      play_emotion.py", "Emotion abspielen (happy, sad, welcoming etc.)"),
    ("      stop_emotion.py", "Emotion stoppen"),
    ("      move_head.py", "Kopf bewegen (left, right, up, down, front)"),
    ("      head_tracking.py", "Kopf-Tracking (Phase 2)"),
    ("      camera.py", "Kamera-Tool (Phase 2)"),
    ("      background_tool_manager.py", "Hintergrund-Tool-Ausführung"),
    ("      task_status.py", "Status von Hintergrund-Tools abfragen"),
    ("      task_cancel.py", "Hintergrund-Tools abbrechen"),
    ("    profiles/", "Profil-Verzeichnis"),
    ("      _reachy_chat_locked_profile/", "Aktives Profil"),
    ("        instructions.txt", "System-Prompt für das LLM"),
    ("        tools.txt", "Liste der aktivierten Tools"),
    ("        sweep_look.py", "Custom Tool: Kopf schwenken"),
    ("        custom_tool.py", "Vorlage für eigene Tools"),
    ("    audio/", "Audio-Module (teilweise entfernt)"),
    ("      speech_tapper.py", "Speech-Tapper (beibehalten)"),
    ("    vision/", "Vision-Module (Phase 2)"),
    ("      processors.py", "Vision-Prozessoren"),
    ("      yolo_head_tracker.py", "YOLO Head-Tracker"),
    ("  pyproject.toml", "Python-Paketdefinition, Dependencies, Entry-Points"),
]

for path, desc in structure:
    c1 = style_cell(ws6, r, 1)
    c1.value = path
    c1.font = code_font
    style_cell(ws6, r, 2).value = desc
    r += 1
r += 1

# ── 3. API-Endpunkte ──
r = doc_section(ws6, r, "3. REST API-Endpunkte (Port 8042)")

endpoints = [
    ("POST /chat", '{"message": "Hello!"}', "text/event-stream",
     "Sendet eine Chat-Nachricht. Antwort wird als SSE-Stream zurückgegeben. "
     "Jede Zeile: 'data: <token>\\n\\n'. Ende: 'data: [DONE]\\n\\n'. "
     "Tool-Aufrufe erscheinen als '[Tool: name]' und '[Result: {...}]'."),
    ("GET /status", "—", "application/json",
     'Gibt {"connected": true/false, "model": "llama3.1:8b"} zurück. '
     "Prüft ob Ollama erreichbar ist."),
    ("GET /history", "—", "application/json",
     "Gibt den Chat-Verlauf als Array zurück: "
     '[{"role": "user", "content": "..."}, {"role": "assistant", "content": "..."}]'),
    ("POST /clear", "—", "application/json",
     "Löscht den Chat-Verlauf. System-Prompt bleibt erhalten. "
     'Gibt {"status": "ok"} zurück.'),
]

headers_api = ["Endpunkt", "Request Body", "Response Type", "Beschreibung"]
for col, h in enumerate(headers_api, 1):
    ws6.cell(row=r, column=col)
# Use 2-column layout for readability
for method, body, resp_type, desc in endpoints:
    r = doc_row(ws6, r, method, f"Body: {body}  |  Response: {resp_type}")
    r = doc_text(ws6, r, f"  {desc}")
r += 1

# ── 4. OllamaChatHandler ──
r = doc_section(ws6, r, "4. OllamaChatHandler — LLM-Backend (ollama_chat.py)")
r = doc_text(ws6, r,
    "Zentrale Klasse für die LLM-Kommunikation. Nutzt die openai-Python-Bibliothek "
    "mit Ollamas OpenAI-kompatibler API (localhost:11434/v1).")
r = doc_row(ws6, r, "Klasse", "OllamaChatHandler", True)
r = doc_row(ws6, r, "Client", "AsyncOpenAI(base_url='http://localhost:11434/v1', api_key='ollama')", True)
r = doc_row(ws6, r, "Modell", "Konfigurierbar via OLLAMA_MODEL (lokal: qwen3:8b-q4km via .env)")
r = doc_row(ws6, r, "Max Tool-Rounds", "5 (verhindert Endlosschleifen bei Tool-Calls)")
r = doc_row(ws6, r, "think-Modus", "Entfernt — verursachte leere Antworten mit qwen3")
r += 1

r = doc_subsection(ws6, r, "Methoden")
methods = [
    ("chat(user_message)", "async generator",
     "Sendet Nachricht an Ollama, yieldet Tokens. Handhabt Tool-Calling-Loops: "
     "wenn LLM ein Tool aufruft, wird es ausgeführt und das Ergebnis zurückgefüttert. "
     "Bis zu 5 Tool-Rounds pro Nachricht."),
    ("check_connection()", "async -> bool",
     "Prüft ob Ollama erreichbar ist (client.models.list())."),
    ("clear_history()", "sync",
     "Löscht Gesprächsverlauf, behält System-Prompt."),
    ("get_history()", "sync -> list",
     "Gibt vereinfachten Verlauf zurück (nur user/assistant Nachrichten)."),
]
for name, ret, desc in methods:
    r = doc_row(ws6, r, name, f"Returns: {ret}")
    r = doc_text(ws6, r, f"  {desc}")
r += 1

# ── 5. Tool-Calling Flow ──
r = doc_section(ws6, r, "5. Tool-Calling — Ablauf")
flow_steps = [
    "1. User sendet Nachricht via POST /chat",
    "2. OllamaChatHandler fügt Nachricht zum Verlauf hinzu",
    "3. Ollama Chat-Completion wird aufgerufen (mit tools-Parameter)",
    "4. Streaming-Antwort wird gelesen:",
    "   a) Text-Tokens → direkt an Browser gestreamt",
    "   b) Tool-Calls → gesammelt bis Stream endet",
    "5. Falls Tool-Calls vorhanden:",
    "   a) Tool wird via dispatch_tool_call() ausgeführt",
    "   b) Ergebnis wird als tool-Message zum Verlauf hinzugefügt",
    "   c) Zurück zu Schritt 3 (LLM reagiert auf Tool-Ergebnis)",
    "6. Falls nur Text → Nachricht wird gespeichert, Stream endet",
]
for step in flow_steps:
    r = doc_text(ws6, r, step)
r += 1

# ── 6. Tool-Referenz ──
r = doc_section(ws6, r, "6. Tool-Referenz — Verfügbare Tools")

tools_doc = [
    ("dance", "dance.py",
     "Spielt einen Tanz ab. Parameter: name (optional, sonst zufällig), repeat (int, default 1). "
     "Nutzt reachy_mini_dances_library. Non-blocking via MovementManager."),
    ("stop_dance", "stop_dance.py",
     "Stoppt den aktuellen Tanz. Keine Parameter."),
    ("play_emotion", "play_emotion.py",
     "Spielt eine aufgenommene Emotion ab. Parameter: emotion (optional, z.B. 'happy', 'sad', "
     "'welcoming1'). Nutzt RecordedMoves aus pollen-robotics/reachy-mini-emotions-library."),
    ("stop_emotion", "stop_emotion.py",
     "Stoppt die aktuelle Emotion. Keine Parameter."),
    ("move_head", "move_head.py",
     "Bewegt den Kopf in eine Richtung. Parameter: direction ('left', 'right', 'up', 'down', 'front')."),
    ("sweep_look", "profiles/.../sweep_look.py",
     "Custom Tool: Schwenkt den Kopf von links nach rechts und zurück. "
     "Rotiert auch den Körper. Dauer ca. 14s. Keine Parameter."),
    ("task_status", "task_status.py",
     "System-Tool: Zeigt Status von Hintergrund-Tools. "
     "Wird automatisch geladen (SystemTool)."),
    ("task_cancel", "task_cancel.py",
     "System-Tool: Bricht einen laufenden Hintergrund-Task ab. "
     "Parameter: task_id. Wird automatisch geladen."),
]

for name, file, desc in tools_doc:
    r = doc_row(ws6, r, name, f"Datei: {file}")
    r = doc_text(ws6, r, f"  {desc}")
r += 1

# ── 7. Konfiguration ──
r = doc_section(ws6, r, "7. Konfiguration")
r = doc_text(ws6, r,
    "Konfiguration erfolgt über Umgebungsvariablen oder .env-Datei im Projektverzeichnis.")

config_vars = [
    ("OLLAMA_BASE_URL", "http://localhost:11434/v1", "Ollama API-Endpunkt"),
    ("OLLAMA_MODEL", "llama3.1:8b", "Modellname für Chat-Completions (lokal: qwen3:8b-q4km via .env)"),
    ("HF_HOME", "./cache", "Hugging Face Cache-Verzeichnis"),
    ("HF_TOKEN", "(leer)", "Hugging Face Token (optional)"),
    ("REACHY_MINI_SKIP_DOTENV", "false", ".env-Datei nicht laden"),
    ("REACHY_MINI_CUSTOM_PROFILE", "(locked)", "Profil-Name (durch LOCKED_PROFILE überschrieben)"),
    ("REACHY_MINI_EXTERNAL_PROFILES_DIRECTORY", "(leer)", "Externes Profil-Verzeichnis"),
    ("REACHY_MINI_EXTERNAL_TOOLS_DIRECTORY", "(leer)", "Externes Tool-Verzeichnis"),
]

r_hdr = r
for col, h in enumerate(["Variable", "Default", "Beschreibung"], 1):
    cell = ws6.cell(row=r_hdr, column=col if col < 3 else 2)
# Use key-value format
for var, default, desc in config_vars:
    r = doc_row(ws6, r, var, f"Default: {default} — {desc}")
r += 1

# ── 8. Profil-System ──
r = doc_section(ws6, r, "8. Profil-System")
r = doc_text(ws6, r,
    "Profile definieren Verhalten und verfügbare Tools des Roboters. "
    "Das aktive Profil ist '_reachy_chat_locked_profile' (in config.py festgelegt).")
r += 1

r = doc_subsection(ws6, r, "instructions.txt — System-Prompt")
r = doc_text(ws6, r,
    "You are Reachy, a small expressive robot made by Pollen Robotics.\n"
    "You can dance, show emotions, move your head, and look around.\n"
    "Be friendly and concise. Use your tools when appropriate.\n"
    "When someone asks you to dance, use the dance tool.\n"
    "When you should express an emotion, use play_emotion.\n"
    "You can look around the room using sweep_look.")
r += 1

r = doc_subsection(ws6, r, "tools.txt — Aktivierte Tools")
r = doc_text(ws6, r, "dance\nstop_dance\nplay_emotion\nstop_emotion\nmove_head\nsweep_look")
r += 1

r = doc_subsection(ws6, r, "Eigene Tools erstellen")
r = doc_text(ws6, r,
    "1. Python-Datei im Profil-Ordner erstellen (z.B. my_tool.py)\n"
    "2. Klasse von Tool erben, name/description/parameters_schema definieren\n"
    "3. async __call__(self, deps: ToolDependencies, **kwargs) implementieren\n"
    "4. Tool-Name in tools.txt eintragen\n"
    "5. App neu starten")
r += 1

# ── 9. MovementManager ──
r = doc_section(ws6, r, "9. MovementManager — Bewegungssystem (moves.py)")
r = doc_text(ws6, r,
    "Der MovementManager steuert alle Roboter-Bewegungen in einem 100Hz Control-Loop. "
    "Er verwaltet primäre Moves (Tanz, Emotionen, Goto) sequenziell und "
    "sekundäre Offsets (Speech, Face-Tracking) additiv.")

mm_details = [
    ("Control-Loop", "100Hz, eigener Thread, monotone Clock"),
    ("Primäre Moves", "Sequenziell via Move-Queue (deque). Nur ein Move gleichzeitig."),
    ("Sekundäre Offsets", "Additiv: Speech-Sway + Face-Tracking. Via Locks thread-safe."),
    ("Idle/Breathing", "Nach 0.3s Inaktivität startet automatisch BreathingMove"),
    ("Listening-Modus", "Friert Antennen ein, blendet sanft zurück beim Aufheben"),
    ("Thread-Safety", "Externe Threads kommunizieren via Command-Queue"),
    ("API: queue_move()", "Move in Queue einreihen"),
    ("API: clear_move_queue()", "Queue leeren, aktuellen Move stoppen"),
    ("API: set_listening()", "Listening-Modus an/aus"),
    ("API: start() / stop()", "Worker-Thread starten/stoppen"),
]

for label, detail in mm_details:
    r = doc_row(ws6, r, label, detail)
r += 1

# ── 10. Browser Chat-UI ──
r = doc_section(ws6, r, "10. Browser Chat-UI (static/)")
r = doc_text(ws6, r,
    "Single-Page Chat-Interface mit Dark Theme. Kommuniziert via REST mit dem FastAPI-Backend.")

ui_features = [
    ("SSE-Streaming", "Antwort-Tokens werden live angezeigt (ReadableStream API)"),
    ("Typing-Indikator", "'Reachy is thinking...' Animation während LLM verarbeitet"),
    ("Tool-Call-Anzeige", "Tool-Aufrufe und Ergebnisse als monospace-Karten im Chat"),
    ("Status-Badge", "Zeigt Verbindungsstatus (connected/disconnected) und Modellname"),
    ("Auto-Resize", "Textarea wächst mit bis max 120px Höhe"),
    ("Tastenkürzel", "Enter = Senden, Shift+Enter = Zeilenumbruch"),
    ("Clear-Button", "Löscht Chat-Verlauf (Frontend + Backend)"),
    ("Responsive", "Funktioniert auf Desktop und Mobilgeräten"),
    ("Status-Polling", "Alle 10 Sekunden wird /status geprüft"),
    ("Nachrichten-Typen", "user (lila), bot (blau), tool (monospace), error (rot)"),
]

for feat, desc in ui_features:
    r = doc_row(ws6, r, feat, desc)
r += 1

# ── 11. Datenfluss ──
r = doc_section(ws6, r, "11. Datenfluss — Ende-zu-Ende")
flow = [
    ("1. Browser", "User tippt Nachricht, klickt Send oder drückt Enter"),
    ("2. main.js", "POST /chat mit JSON body, öffnet ReadableStream"),
    ("3. FastAPI (main.py)", "Empfängt Request, ruft chat_handler.chat() auf"),
    ("4. ollama_chat.py", "Fügt Nachricht zum Verlauf, ruft Ollama API auf"),
    ("5. Ollama (11434)", "Verarbeitet mit llama3.1:8b, streamt Tokens/Tool-Calls"),
    ("6. ollama_chat.py", "Falls Tool-Call: dispatch via core_tools.py"),
    ("7. tools/*.py", "Tool wird ausgeführt (z.B. dance → MovementManager.queue_move)"),
    ("8. MovementManager", "Move wird in 100Hz-Loop ausgeführt → set_target() an SDK"),
    ("9. reachy-mini SDK", "Sendet Befehle via WebSocket an Daemon"),
    ("10. Daemon (8000)", "Steuert MuJoCo-Simulation (oder Hardware)"),
    ("11. ollama_chat.py", "Tool-Ergebnis zurück an Ollama, LLM generiert finale Antwort"),
    ("12. main.py", "SSE-Stream: 'data: <token>\\n\\n' pro Token"),
    ("13. main.js", "Tokens werden live im Chat-Bubble angezeigt"),
]

for step, desc in flow:
    r = doc_row(ws6, r, step, desc)
r += 1

# ── 12. Bekannte Einschränkungen ──
r = doc_section(ws6, r, "12. Bekannte Einschränkungen (Phase 1)")
limitations = [
    ("Voice (PipeWire)", "Push-to-Talk implementiert. Bekanntes Problem: sd.InputStream bei 16 kHz liefert Stille via PipeWire. Fix vorhanden (Resampling) — Verifikation ausstehend."),
    ("Keine Kamera", "Simulation hat keine Kamera. Phase 2 aktiviert CameraWorker."),
    ("CPU-Latenz", "Ollama auf CPU: 2-5s für erste Tokens. Mit GPU deutlich schneller."),
    ("Tool-Call-Qualität", "Hängt vom Modell ab. qwen3:8b-q4km funktioniert gut."),
    ("Kein Persistenter Chat", "Verlauf geht bei App-Neustart verloren."),
    ("Single-User", "Kein Multi-User-Support. Ein Chat-Verlauf pro App-Instanz."),
]

for limit, desc in limitations:
    r = doc_row(ws6, r, limit, desc)
r += 1

# ── 13. Voice-System ──
r = doc_section(ws6, r, "13. Voice-System (voice/)")
r = doc_text(ws6, r,
    "Push-to-Talk Sprachsteuerung über den Browser. "
    "WebSocket-Endpoint /ws/voice. Läuft serverseitig, komplett lokal, kein Cloud-Dienst.")

r = doc_subsection(ws6, r, "Komponenten")
voice_components = [
    ("Spracherkennung (STT)", "faster-whisper (small-Modell, CPU, Deutsch)"),
    ("Sprachausgabe (TTS)", "Piper — de_DE-thorsten-high.onnx (ONNX, lokal)"),
    ("Stille-Erkennung (VAD)", "Silero VAD — erkennt Sprachpausen automatisch"),
    ("Aufnahme", "sounddevice — native Geräterate (44 kHz), Resampling auf 16 kHz"),
    ("Pipeline", "voice/pipeline.py — async: Mikrofon → STT → Ollama → TTS → Lautsprecher"),
    ("Modell-Download", "de_DE-thorsten-high.onnx wird beim ersten Start via HuggingFace geladen"),
]
for comp, detail in voice_components:
    r = doc_row(ws6, r, comp, detail)
r += 1

r = doc_subsection(ws6, r, "WebSocket-Protokoll (Browser ↔ Server)")
ws_protocol = [
    ('{"type": "start_recording"}', "Mikrofon öffnen, Aufnahme startet"),
    ('{"type": "stop_recording"}', "Aufnahme stoppen → Whisper STT läuft → Antwort: transcript"),
    ('{"type": "transcript", "text": "..."}', "Server → Browser: erkannter Text"),
    ('{"type": "speak", "text": "..."}', "Browser → Server: Text vorlesen lassen"),
    ('{"type": "status", "value": "listening|processing|speaking|null"}', "Serverstatus-Updates"),
]
for msg, desc in ws_protocol:
    r = doc_row(ws6, r, msg, desc, is_code=True)
r += 1

r = doc_subsection(ws6, r, "Bekanntes Problem & Fix")
r = doc_text(ws6, r,
    "Problem: sd.InputStream bei 16 kHz liefert via PipeWire Stille.\n"
    "Fix in vad.py: Aufnahme bei nativer Geräterate (44 kHz), anschließend Resampling auf 16 kHz.\n"
    "Status: Implementiert — Verifikation nach App-Neustart ausstehend.")
r += 1

r = doc_subsection(ws6, r, "Standalone-Test")
r = doc_row(ws6, r, "Befehl",
    "/home/sbin/reachy/.venv/bin/python3 /home/sbin/reachy/voice/pipeline.py", is_code=True)
r = doc_text(ws6, r, "Startet die komplette Pipeline ohne Browser (Mikrofon → STT → Ollama → TTS).")

# Save
output_path = f"/home/sbin/reachy/reachy_chat_status_{date.today().strftime('%Y-%m-%d')}.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
